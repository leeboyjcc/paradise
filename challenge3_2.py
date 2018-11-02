# -*- coding: utf-8 -*-

from openpyxl import load_workbook, Workbook
from datetime import datetime

filename='/home/shiyanlou/Code/courses.xlsx'
def combine():
    wb = load_workbook(filename)
    ws1 = wb['students']
    ws2 = wb['time']

    ws3 = wb.create_sheet('combine')

    r = ws1.max_row
    c = ws1.max_column

    for i in range(1,r+1):
        for j in range(1,c+2):
            if j == c+1:
                ws3.cell(row=i,column=j).value = ws2.cell(row=i,column=j-1).value
            else:
                ws3.cell(row=i,column=j).value = ws1.cell(row=i,column=j).value

    wb.save(filename)

def split():
    wb = load_workbook(filename)
    ws = wb['combine']

    year = set()
    for row in ws.iter_rows(min_row=2,max_row=ws.max_row): #过滤掉首行标题
        year.add(str(row[0].value)[:4])

    for y in year:
        fn = '/home/shiyanlou/Code/{}.xlsx'.format(y)
        wb_year = Workbook()
        ws_year = wb_year.active
        ws_year.title = y
        
        for row in ws_year.iter_rows(min_row=1, max_row=1):
            ws_year.append([cell.value for cell in row]) # 直接添加list

        for row in ws_year.iter_rows(min_row=2, max_row=ws.max_row):
            if str(row[0].value)[:4] == y:
                    ws_year.append([cell.value for cell in row])

        wb_year.save(fn)


if __name__=='__main__':
    combine()
    split()
